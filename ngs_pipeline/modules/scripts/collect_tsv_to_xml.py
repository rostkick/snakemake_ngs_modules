import pandas as pd
import os
import sys
import subprocess
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

log_file = snakemake.log[0]
input_files = snakemake.input.tsv
output_file = snakemake.output.xlsx
params_file = snakemake.params.tsv
hs_metrics_dir = snakemake.params.hs_metrics_dir
vcf_dir = snakemake.params.vcf_dir
bcftools = snakemake.params.bcftools


def get_hs_metrics(sample, hs_metrics_dir):
    path = os.path.join(hs_metrics_dir, f"{sample}.hs_metrics.tsv")
    if not os.path.exists(path):
        print(f"[QC] hs_metrics not found: {path}")
        return {}
    try:
        header = None
        values = None
        with open(path) as f:
            for line in f:
                line = line.rstrip('\n')
                if line.startswith('#') or line.strip() == '':
                    continue
                if header is None:
                    header = line.split('\t')
                elif values is None:
                    values = line.split('\t')
                    break
        if header is None or values is None:
            print(f"[QC] hs_metrics parse failed for {sample}: header or values missing")
            return {}
        d = dict(zip(header, values))
        return {
            'Mean_Coverage':    float(d.get('MEAN_TARGET_COVERAGE', 'nan')),
            'PCT_10X':          float(d.get('PCT_TARGET_BASES_10X', 'nan')),
            'PCT_20X':          float(d.get('PCT_TARGET_BASES_20X', 'nan')),
            'PCT_30X':          float(d.get('PCT_TARGET_BASES_30X', 'nan')),
            'Fold_Enrichment':  float(d.get('FOLD_ENRICHMENT', 'nan')),
            'Zero_Cvg_PCT':     float(d.get('ZERO_CVG_TARGETS_PCT', 'nan')),
            'PCT_Off_Bait':     float(d.get('PCT_OFF_BAIT', 'nan')),
        }
    except Exception as e:
        print(f"[QC] hs_metrics error for {sample}: {e}")
        return {}


def get_vcf_stats(sample, vcf_dir, bcftools):
    vcf = os.path.join(vcf_dir, f"{sample}.joint.annotated.vcf.gz")
    if not os.path.exists(vcf):
        return {}
    try:
        # SNP, indel, Ti/Tv
        cmd = f"{bcftools} view -i 'GT!=\"0/0\" && GT!=\"./.\"' {vcf} | {bcftools} stats"
        out = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL).decode()
        snp = indel = titv = ins = dels = None
        for line in out.splitlines():
            if 'number of SNPs:' in line:
                snp = int(line.strip().split()[-1])
            elif 'number of indels:' in line:
                indel = int(line.strip().split()[-1])
            elif line.startswith('TSTV'):
                titv = float(line.strip().split()[4])
            elif line.startswith('IDD'):
                parts = line.strip().split()
                size = int(parts[2])
                count = int(parts[3])
                if size > 0:
                    ins = (ins or 0) + count
                elif size < 0:
                    dels = (dels or 0) + count
        ins_del = round(ins / dels, 3) if ins and dels else None

        # het/hom + mean GQ/DP
        cmd2 = f"{bcftools} query -f '[%GT\\t%GQ\\t%DP]\\n' {vcf}"
        out2 = subprocess.check_output(cmd2, shell=True, stderr=subprocess.DEVNULL).decode()
        het = hom = 0
        het_gq_sum = het_dp_sum = hom_gq_sum = hom_dp_sum = 0
        het_gq_n = het_dp_n = hom_gq_n = hom_dp_n = 0
        for line in out2.splitlines():
            parts = line.strip().split('\t')
            if len(parts) != 3:
                continue
            gt, gq, dp = parts
            if gt == '0/1':
                het += 1
                if gq not in ('.', ''):
                    het_gq_sum += float(gq); het_gq_n += 1
                if dp not in ('.', ''):
                    het_dp_sum += float(dp); het_dp_n += 1
            elif gt == '1/1':
                hom += 1
                if gq not in ('.', ''):
                    hom_gq_sum += float(gq); hom_gq_n += 1
                if dp not in ('.', ''):
                    hom_dp_sum += float(dp); hom_dp_n += 1
        het_hom      = round(het / hom, 3) if hom > 0 else None
        mean_gq_het  = round(het_gq_sum / het_gq_n, 2) if het_gq_n > 0 else None
        mean_dp_het  = round(het_dp_sum / het_dp_n, 2) if het_dp_n > 0 else None
        mean_gq_hom  = round(hom_gq_sum / hom_gq_n, 2) if hom_gq_n > 0 else None
        mean_dp_hom  = round(hom_dp_sum / hom_dp_n, 2) if hom_dp_n > 0 else None

        # sex from chrX
        cmd3 = f"{bcftools} query -r chrX -f '[%GT]\\n' {vcf}"
        out3 = subprocess.check_output(cmd3, shell=True, stderr=subprocess.DEVNULL).decode()
        x_het = x_hom = 0
        for gt in out3.splitlines():
            if gt == '0/1':
                x_het += 1
            elif gt == '1/1':
                x_hom += 1
        ratio = x_het / (x_hom + 0.001)
        sex = 'F' if ratio > 0.3 else 'M'

        return {
            'SNP':          snp,
            'Indels':       indel,
            'Ins/Del':      ins_del,
            'Ti/Tv':        titv,
            'Het':          het,
            'Hom':          hom,
            'Het/Hom':      het_hom,
            'Mean_GQ_het':  mean_gq_het,
            'Mean_DP_het':  mean_dp_het,
            'Mean_GQ_hom':  mean_gq_hom,
            'Mean_DP_hom':  mean_dp_hom,
            'Sex':          sex,
        }
    except Exception as e:
        print(f"[QC] vcf stats error for {sample}: {e}")
        return {}


def build_qc_df(sheet_data, hs_metrics_dir, vcf_dir, bcftools):
    rows = []
    for data in sheet_data:
        sample = data['sample_short']
        row = {'Sample': sample, 'Full_Name': data['full_name']}
        row.update(get_hs_metrics(sample, hs_metrics_dir))
        row.update(get_vcf_stats(sample, vcf_dir, bcftools))
        rows.append(row)
    cols = ['Sample', 'Full_Name',
            'Mean_Coverage', 'PCT_10X', 'PCT_20X', 'PCT_30X',
            'Fold_Enrichment', 'Zero_Cvg_PCT', 'PCT_Off_Bait',
            'SNP', 'Indels', 'Ins/Del', 'Ti/Tv',
            'Het', 'Hom', 'Het/Hom',
            'Mean_GQ_het', 'Mean_DP_het',
            'Mean_GQ_hom', 'Mean_DP_hom',
            'Sex']
    return pd.DataFrame(rows, columns=[c for c in cols if c in pd.DataFrame(rows).columns])


def add_qc_sheet(workbook, qc_df):
    ws = workbook.create_sheet('QC_Summary', 1)  # after Legend

    # back-link to Legend
    back_cell = ws.cell(row=1, column=1)
    back_cell.value = '← Legend'
    hl = Hyperlink(ref='', location="'Legend'!A1")
    hl.display = '← Legend'
    back_cell.hyperlink = hl
    back_cell.font = Font(color='0000FF', underline='single', bold=True)

    # header row
    header_fill = PatternFill('solid', start_color='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx, col_name in enumerate(qc_df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # data rows
    for row_idx, row in enumerate(qc_df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    # auto column width
    for col_idx, col_name in enumerate(qc_df.columns, start=1):
        max_len = max(len(str(col_name)),
                      max((len(str(v)) for v in qc_df.iloc[:, col_idx-1]), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 25)


def merge_tsv_with_legend(input_files, output_file, params_df):
    sheet_data = []
    for file in input_files:
        sample_short = os.path.basename(file).split('.')[0]
        full_name = params_df.loc[params_df['sample'] == sample_short, 'full_name'].values[0]
        sheet_data.append({
            'file': file,
            'sample_short': sample_short,
            'full_name': full_name
        })

    print(f"[QC] Collecting QC metrics for {len(sheet_data)} samples...")
    qc_df = build_qc_df(sheet_data, hs_metrics_dir, vcf_dir, bcftools)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Legend: write with startrow=2 to leave rows 1 for QC link, 2 for headers
        legend_data = [[d['full_name'], d['sample_short']] for d in sheet_data]
        legend_df = pd.DataFrame(legend_data, columns=['Sheet Name', 'Sample'])
        legend_df.to_excel(writer, sheet_name='Legend', index=False, startrow=2)

        for data in sheet_data:
            df = pd.read_csv(data['file'], sep='\t')
            df.to_excel(writer, sheet_name=data['full_name'], index=False,
                        float_format='%.8f', startrow=1)

        workbook = writer.book

        # QC sheet (after Legend)
        add_qc_sheet(workbook, qc_df)

        # --- Legend sheet ---
        legend_ws = workbook['Legend']

        # Row 1: link to QC_Summary
        qc_cell = legend_ws.cell(row=1, column=1)
        qc_cell.value = '→ QC_Summary'
        hl = Hyperlink(ref='', location="'QC_Summary'!A1")
        hl.display = '→ QC_Summary'
        qc_cell.hyperlink = hl
        qc_cell.font = Font(color='0000FF', underline='single', bold=True)

        # Row 3 (headers written by pandas at startrow=2, i.e. Excel row 3): bold
        for cell in legend_ws[3]:
            cell.font = cell.font.copy(bold=True)

        # Rows 4+: hyperlinks to sample sheets
        for i, data in enumerate(sheet_data, start=4):
            sheet_name = data['full_name']
            cell = legend_ws.cell(row=i, column=1)
            hl = Hyperlink(ref='', location=f"'{sheet_name}'!A1")
            hl.display = sheet_name
            cell.hyperlink = hl
            cell.value = sheet_name
            cell.font = cell.font.copy(color='0000FF', underline='single')

        # back-links from sample sheets
        for data in sheet_data:
            sheet_name = data['full_name']
            ws = workbook[sheet_name]
            back_cell = ws.cell(row=1, column=1)
            back_cell.value = '← Legend'
            hl = Hyperlink(ref='', location="'Legend'!A1")
            hl.display = '← Legend'
            back_cell.hyperlink = hl
            back_cell.font = Font(color='0000FF', underline='single', bold=True)


with open(log_file, 'w') as f:
    sys.stderr = sys.stdout = f
    params_df = pd.read_csv(params_file, sep='\t')
    merge_tsv_with_legend(input_files, output_file, params_df)